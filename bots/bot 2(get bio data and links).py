from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time
import chromedriver_autoinstaller
import random
import openpyxl


def sleeper():
    time.sleep(random.randrange(3, 5))


insta_accounts = []
insta_accounts_file = open("..\\resources\\insta accounts.txt", "r")
for i in insta_accounts_file:
    j = i
    if i[-1] == "\n":
        j = j[:-1]
    insta_accounts.append(j.split(" "))

# Logging into Instagram
for account in insta_accounts:
    chromedriver_autoinstaller.install()
    browser = webdriver.Chrome()
    browser.implicitly_wait(1)
    sleeper()
    browser.get('https://www.instagram.com/')
    sleeper()
    username_input = browser.find_element(By.CSS_SELECTOR, "input[name='username']")
    password_input = browser.find_element(By.CSS_SELECTOR, "input[name='password']")
    username_input.send_keys(account[0])
    password_input.send_keys(account[1])
    login_button = browser.find_element(By.XPATH, "//button[@type='submit']")
    login_button.click()
    sleeper()

    # Opening files
    user_following_file = open("..\\resources\\users' following list.txt", "r")
    done_with = open("..\\resources\\done with users' following list.txt", "r+")

    # Opening Excel file and creating a sheet
    book = openpyxl.load_workbook("..\\resources\\Bio data and links.xlsx")
    sheet = book["Sheet1"]

    brands = []
    user_count = 0

    # making user's following list
    user_following_list = []
    for i in user_following_file:
        if i[-1] == "\n":
            user_following_list.append(i[:-1])
        else:
            user_following_list.append(i)
    if not user_following_list:
        print("No usernames found")
        exit()

    # making done with user's following list
    done_with_user_following_list = []
    for i in done_with:
        if i[-1] == "\n":
            done_with_user_following_list.append(i[:-1])
        else:
            done_with_user_following_list.append(i)
    done_with.read()

    # # clearing done with file
    # done_with.close()
    # open("..\\resources\\done with users' following list.txt", "w").close()
    # done_with = open("..\\resources\\done with users' following list.txt", "r+")

    # row and column set
    row = 2
    while True:
        if not sheet.cell(row, 1).value:
            break
        row += 1
    column = 1

    # iterating following list
    for i in user_following_list:
        if user_count == 200:
            break
        if i not in done_with_user_following_list and i != "Verified":
            print(i)
            brands = []
            browser.get("https://www.instagram.com/" + i + "/")
            sleeper()
            sleeper()
            try:
                bio_data = browser.find_element(By.CLASS_NAME, "_aa_c")
            except NoSuchElementException:
                continue
            sheet.cell(row, column).value = i
            sheet.cell(row, column + 1).value = "https://www.instagram.com/" + i + "/"
            sheet.cell(row, column + 2).value = bio_data.text
            bio_data_tokens = bio_data.text.split()

            links = ""
            mentions = ""
            count_1 = 0
            count_2 = 0

            # iterate bio data tokens and find links and mentions
            for j in range(len(bio_data_tokens)):
                if "@" in bio_data_tokens[j] and bio_data_tokens[j].index("@") == 0:
                    mentions += str(bio_data_tokens[j])
                    mentions += "\n"
                    count_1 = 1
                if "." in bio_data_tokens[j] and "/" in bio_data_tokens[j] or bio_data_tokens[j].endswith(".com") or\
                        bio_data_tokens[j].endswith(".in") or bio_data_tokens[j].endswith(".org") or bio_data_tokens[j]\
                        .endswith(".net") or bio_data_tokens[j].endswith(".info") or \
                        bio_data_tokens[j].endswith(".biz"):
                    links += str(bio_data_tokens[j])
                    links += "\n"
                    count_2 = 1
            links = links[:-1]
            mentions = mentions[:-1]

            sheet.cell(row, column + 3).value = mentions
            sheet.cell(row, column + 4).value = links
            links_list = []
            if links.count("\n") == 0:
                links_list.append(links)
            else:
                links_list = links.split("\n")
            for j in links_list:
                if "linktr" in j:
                    sheet.cell(row, column + 5).value = j

                    if "http" in j:
                        browser.get(j)
                    else:
                        browser.get('https://' + j)
                    sleeper()

                    links = browser.find_elements(By.TAG_NAME, "a")
                    linktree_links = ""

                    for link in links:
                        linki = link.get_attribute("href")
                        if not linki:
                            continue
                        if "instagram" not in linki and "tiktok" not in linki and "spotify" not in linki and "podcast" \
                                not in linki and "linktr.ee" not in linki and "facebook" not in linki and "mailto" \
                                not in linki and "amazon" not in linki and "amzn.to" not in linki \
                                and "gmail" not in linki and "fb" not in linki and "pinterest" not in linki \
                                and "youtube" not in linki and "youtu" not in linki:
                            linkj = linki[linki.index("//") + 2:]
                            linktree_links += linkj[:linkj.index("/")] + "(" + linki + ")" + "\n"

                    linktree_links = linktree_links[:-1]
                    sheet.cell(row, column + 6).value = linktree_links
                if "beacons.ai" in j:
                    sheet.cell(row, column + 7).value = j

                    if "http" in j:
                        browser.get(j)
                    else:
                        browser.get('https://' + j)
                    sleeper()

                    links = browser.find_elements(By.TAG_NAME, "a")
                    linktree_links = ""

                    for link in links:
                        linki = link.get_attribute("href")
                        if not linki:
                            continue
                        if "instagram" not in linki and "tiktok" not in linki and "spotify" not in linki and "podcasts" \
                                not in linki and "beacons.ai" not in linki and "facebook" not in linki and "ngl.link" not \
                                in linki and "mailto" not in linki and "amazon" not in linki and "amzn.to" not in linki \
                                and "@gmail.com" not in linki and "fb.watch" not in linki and "pinterest" not in linki \
                                and "unfold.onelink.me" not in linki and "youtube" not in linki \
                                and "youtu" not in linki:
                            linkj = linki[linki.index("//") + 2:]
                            linktree_links += linkj[:linkj.index("/")] + "(" + linki + ")" + "\n"

                    linktree_links = linktree_links[:-1]
                    sheet.cell(row, column + 8).value = linktree_links
                if "msha.ke" in j:
                    sheet.cell(row, column + 9).value = j

                    if "http" in j:
                        browser.get(j)
                    else:
                        browser.get('https://' + j)
                    sleeper()

                    links = browser.find_elements(By.TAG_NAME, "a")
                    linktree_links = ""

                    for link in links:
                        linki = link.get_attribute("href")
                        if not linki:
                            continue
                        if "instagram" not in linki and "tiktok" not in linki and "spotify" not in linki and "podcasts" \
                                not in linki and "milkshake.app" not in linki and "facebook" not in linki and "ngl.link" not \
                                in linki and "mailto" not in linki and "amazon" not in linki and "amzn.to" not in linki \
                                and "@gmail.com" not in linki and "fb.watch" not in linki and "pinterest" not in linki \
                                and "unfold.onelink.me" not in linki and "youtube" not in linki \
                                and "youtu" not in linki:
                            linkj = linki[linki.index("//") + 2:]
                            linktree_links += linkj[:linkj.index("/")] + "(" + linki + ")" + "\n"

                    linktree_links = linktree_links[:-1]
                    sheet.cell(row, column + 10).value = linktree_links
                if "bio.site" in j:
                    sheet.cell(row, column + 11).value = j

                    if "http" in j:
                        browser.get(j)
                    else:
                        browser.get('https://' + j)
                    sleeper()

                    links = browser.find_elements(By.TAG_NAME, "a")
                    linktree_links = ""

                    for link in links:
                        linki = link.get_attribute("href")
                        if not linki:
                            continue
                        if "instagram" not in linki and "tiktok" not in linki and "spotify" not in linki and "podcasts" \
                                not in linki and "bio.site" not in linki and "facebook" not in linki and "ngl.link" not \
                                in linki and "mailto" not in linki and "amazon" not in linki and "amzn.to" not in linki \
                                and "@gmail.com" not in linki and "fb.watch" not in linki and "pinterest" not in linki \
                                and "unfold.onelink.me" not in linki and "youtube" not in linki \
                                and "youtu" not in linki:
                            linkj = linki[linki.index("//") + 2:]
                            linktree_links += linkj[:linkj.index("/")] + "(" + linki + ")" + "\n"

                    linktree_links = linktree_links[:-1]
                    sheet.cell(row, column + 12).value = linktree_links
                if "zez.am" in j:
                    sheet.cell(row, column + 13).value = j

                    if "http" in j:
                        browser.get(j)
                    else:
                        browser.get('https://' + j)
                    sleeper()

                    links = browser.find_elements(By.TAG_NAME, "a")
                    linktree_links = ""

                    for link in links:
                        linki = link.get_attribute("href")
                        if not linki:
                            continue
                        if "instagram" not in linki and "tiktok" not in linki and "spotify" not in linki and "podcasts"\
                                not in linki and "zezam.io" not in linki and "facebook" not in linki and "ngl.link" not\
                                in linki and "mailto" not in linki and "amazon" not in linki and "amzn.to" not in linki\
                                and "@gmail.com" not in linki and "fb.watch" not in linki and "pinterest" not in linki\
                                and "unfold.onelink.me" not in linki and "youtube" not in linki \
                                and "youtu" not in linki:
                            linkj = linki[linki.index("//") + 2:]
                            linktree_links += linkj[:linkj.index("/")] + "(" + linki + ")" + "\n"

                    linktree_links = linktree_links[:-1]
                    sheet.cell(row, column + 14).value = linktree_links

            book.save("..\\resources\\Bio data and links.xlsx")
            done_with.write(i + "\n")
            done_with_user_following_list.append(i)
            user_count += 1
            row += 1

    user_following_file.close()
    done_with.close()

    book.save("..\\resources\\Bio data and links.xlsx")
    book.close()
