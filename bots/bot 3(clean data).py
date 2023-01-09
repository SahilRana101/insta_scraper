import openpyxl

# Opening Excel file and creating a sheet
book = openpyxl.load_workbook("..\\resources\\Bio data and links.xlsx")
sheet = book["Sheet1"]

brands = []
val = []

# row and column set
row = 2
while True:
    if not sheet.cell(row, 1).value:
        break
    row += 1
column = 1

u_row = 2
while True:
    if not sheet.cell(u_row, 16).value:
        break
    u_row += 1
u_column = 16

for i in range(2, u_row):
    a = sheet.cell(i, u_column).value
    b = a[:a.index("=") - 1]
    a = a[a.index("=") + 2:]
    brands.append(a)
    val.append(b)

keywords = ["amazon", "etsy", "garmin", "pulte", "facebook", "instagram", "amzn", "airbnb", "tiktok", "youtu",
            "buymeacoffee", "twitter", "pinterest", "linkedin", "paypal", "target", "venmo", "patreon", "calendly",
            "photograph", "canva", "ebook", "fundraise", "spotify", "podcast", "gmail", "fb", "abnb"]

popper = []

for i in range(len(brands)):
    for j in keywords:
        if j in brands[i]:
            popper.append(i)

for i in reversed(popper):
    brands.pop(i)
    val.pop(i)

for i in range(2, row + 1):
    for j in keywords:
        if sheet.cell(i, 7).value:
            if j in sheet.cell(i, 7).value:
                a = sheet.cell(i, 7).value
                b = a.split("\n")
                print(len(b))
                for k in b:
                    if j in k:
                        b.pop(b.index(k))
                        break
                a = ""
                for k in b:
                    a += k + "\n"
                a = a[:-1]
                if a == "":
                    sheet.cell(i, 7).value = None
                else:
                    sheet.cell(i, 7).value = a
        if sheet.cell(i, 5).value:
            if j in sheet.cell(i, 5).value:
                a = sheet.cell(i, 5).value
                b = a.split("\n")
                print(len(b))
                for k in b:
                    if j in k:
                        b.pop(b.index(k))
                        break
                a = ""
                if not (b == []):
                    for k in b:
                        a += k + "\n"
                    a = a[:-1]
                    if a == "":
                        sheet.cell(i, 5).value = None
                    else:
                        sheet.cell(i, 5).value = a
                else:
                    sheet.cell(i, 5).value = None
        if sheet.cell(i, 9).value:
            if j in sheet.cell(i, 9).value:
                a = sheet.cell(i, 9).value
                b = a.split("\n")
                print(len(b))
                for k in b:
                    if j in k:
                        b.pop(b.index(k))
                        break
                a = ""
                if not (b == []):
                    for k in b:
                        a += k + "\n"
                    a = a[:-1]
                    if a == "":
                        sheet.cell(i, 9).value = None
                    else:
                        sheet.cell(i, 9).value = a
                else:
                    sheet.cell(i, 9).value = None
        if sheet.cell(i, 11).value:
            if j in sheet.cell(i, 11).value:
                a = sheet.cell(i, 11).value
                b = a.split("\n")
                print(len(b))
                for k in b:
                    if j in k:
                        b.pop(b.index(k))
                        break
                a = ""
                if not (b == []):
                    for k in b:
                        a += k + "\n"
                    a = a[:-1]
                    if a == "":
                        sheet.cell(i, 11).value = None
                    else:
                        sheet.cell(i, 11).value = a
                else:
                    sheet.cell(i, 11).value = None
        if sheet.cell(i, 13).value:
            if j in sheet.cell(i, 13).value:
                a = sheet.cell(i, 13).value
                b = a.split("\n")
                print(len(b))
                for k in b:
                    if j in k:
                        b.pop(b.index(k))
                        break
                a = ""
                if not (b == []):
                    for k in b:
                        a += k + "\n"
                    a = a[:-1]
                    if a == "":
                        sheet.cell(i, 13).value = None
                    else:
                        sheet.cell(i, 13).value = a
                else:
                    sheet.cell(i, 13).value = None
        if sheet.cell(i, 15).value:
            if j in sheet.cell(i, 15).value:
                a = sheet.cell(i, 15).value
                b = a.split("\n")
                print(len(b))
                for k in b:
                    if j in k:
                        b.pop(b.index(k))
                        break
                a = ""
                if not (b == []):
                    for k in b:
                        a += k + "\n"
                    a = a[:-1]
                    if a == "":
                        sheet.cell(i, 15).value = None
                    else:
                        sheet.cell(i, 15).value = a
                else:
                    sheet.cell(i, 15).value = None

for i in range(2, row + 1):
    if not sheet.cell(i, 5).value:
        continue
    else:
        a = sheet.cell(i, 5).value
        if "\n" in a:
            b = a.split("\n")
            print(len(b))
            for k in b:
                if "@" in k and ".com" in k:
                    b.pop(b.index(k))
                    break
            a = ""
            if not (b == []):
                for k in b:
                    a += k + "\n"
                a = a[:-1]
                if a == "":
                    sheet.cell(i, 5).value = None
                else:
                    sheet.cell(i, 5).value = a
            else:
                sheet.cell(i, 5).value = None
        else:
            if "@" in a and ".com" in a:
                sheet.cell(i, 5).value = None

popper = []
for i in range(2, row + 1):
    if not sheet.cell(i, 4).value and not sheet.cell(i, 5).value:
        popper.append(i)

x = sheet.cell(u_row + 1, u_column).value

for i in reversed(popper):
    sheet.delete_rows(i, 1)

sheet.delete_cols(16, 1)
sheet.cell(1, 16).value = "Ranked links"

for i in range(len(brands)):
    sheet.cell(i + 2, 16).value = val[i] + " = " + brands[i]

sheet.cell(u_row + 1, u_column).value = x

book.save("..\\resources\\Bio data and links.xlsx")
book.close()
