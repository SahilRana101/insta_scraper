import openpyxl


# Opening Excel file and creating a sheet
book = openpyxl.load_workbook("..\\resources\\Bio data and links.xlsx")
sheet = book["Sheet1"]

brands = []

# row and column set
row = 2
while True:
    if not sheet.cell(row, 1).value:
        break
    row += 1
column = 1

u_row = 2
while True:
    if not sheet.cell(u_row, 16).value:
        break
    u_row += 1
u_column = 16

if u_row == 2:
    x = 2
else:
    x = int(sheet.cell(u_row + 1, u_column).value)

for i in range(x, row+1):
    for j in [5, 7, 9, 11, 13, 15]:
        x = sheet.cell(i, j).value
        if not x:
            continue
        else:
            for k in x.split("\n"):
                if "linktr" not in k and "bio.site" not in k and "beacons.ai" not in k and "msha.ke" not in k and\
                        "zez.am" not in k:
                    brands.append(k)

if not brands == []:
    rank_brands_count = []
    rank_brands_names = []
    for j in range(2, u_row):
        a = sheet.cell(j, u_column).value
        rank_brands_count.append(int(a[:a.index("=") - 1]))
        rank_brands_names.append(a[a.index("=") + 2:])
    for j in brands:
        b = j[j.index("//") + 2:]
        b = b[:b.index("/")]
        if b not in rank_brands_names:
            rank_brands_names.append(b)
            rank_brands_count.append(1)
        else:
            rank_brands_count[rank_brands_names.index(b)] += 1

    for j in range(2, len(rank_brands_names) + 2):
        max_count = max(rank_brands_count)
        c = rank_brands_names[rank_brands_count.index(max_count)]
        sheet.cell(j, u_column).value = str(max_count) + " = " + c
        rank_brands_count[rank_brands_count.index(max_count)] = 0

    sheet.cell(len(rank_brands_names) + 3, u_column).value = row - 1

book.save("..\\resources\\Bio data and links.xlsx")
book.close()
